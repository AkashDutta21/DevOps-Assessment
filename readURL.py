def readUrl(ASSESSMENT_FILE, ACTIVE_SHEET):

    from openpyxl import load_workbook

    # declare all required variables
    sheet = ACTIVE_SHEET
    assessment_file = ASSESSMENT_FILE

    # Load the Excel workbook
    workbook = load_workbook(assessment_file)

    # Select the specific worksheet by name or index
    worksheet = workbook[sheet]  # Provide Sheet Data

    row_num = 2
    col_num = 2

    cell_value = worksheet['B2'].value

    return cell_value

# # Iterate over rows and access cell values
# for row in worksheet.iter_rows(min_row=1, values_only=True):
#     cell_value = row[0]  # Accessing the first column, adjust as needed
#     print(cell_value)


